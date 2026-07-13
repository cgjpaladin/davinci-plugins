"""
批量命名工具 · WebView 生产版
Python 后端 + HTML/CSS 前端
"""
import os, sys, json, re, shutil, statistics
from datetime import datetime
from urllib.parse import unquote

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from shared.naming import (
    FIELD_CONFIG, DISPLAY_FIELDS, METHOD_DESC_MAP, FIELD_RULES,
    build_filename, parse_filename, build_folder,
    MEDIA_EXT, sanitize_text,
)
from shared.naming_checks import check_zero_byte, check_double_ext, check_size_anomaly

import webview
import logging
_log = logging.getLogger("renamer_web")
_log.setLevel(logging.DEBUG)
try:
    if sys.platform == "win32":
        _log_dir = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")),
                                "批量命名工具", "Logs")
    else:
        _log_dir = os.path.join(os.path.expanduser("~"), "Library", "Logs", "批量命名工具")
    os.makedirs(_log_dir, exist_ok=True)
    _hdlr = logging.FileHandler(os.path.join(_log_dir, "renamer.log"))
    _hdlr.setFormatter(logging.Formatter('%(asctime)s %(message)s'))
    _log.addHandler(_hdlr)
except Exception:
    pass  # 日志文件不可用时静默跳过

# pyinstaller 打包后 sys._MEIPASS = app bundle 资源目录
_BASE_DIR = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
_SHARED_DIR = os.path.join(_BASE_DIR, 'shared')
if os.path.isdir(_SHARED_DIR) and _SHARED_DIR not in sys.path:
    sys.path.insert(0, os.path.dirname(_SHARED_DIR))

# 视频扩展名集合 + 文件对话框过滤器
SUPPORTED_EXT = {'.mp4','.mov','.mxf','.avi','.mkv','.webm','.m4v','.mts','.mpg','.mpeg','.wmv','.3gp','.flv','.r3d','.braw',
                  '.jpg','.jpeg','.png','.bmp','.tiff','.tif','.gif','.webp','.tga','.targa','.psd'}
_DIALOG_FILTER = "媒体文件 (" + ";".join(sorted(e.replace(".","*.") for e in SUPPORTED_EXT)) + ")"

CFG_FILE = os.path.join(os.path.expanduser("~"), ".renamer_saved.json")
PATH_RE = re.compile(r"^(\d{8})_(.+)$")

_saved_defaults = {}
if os.path.exists(CFG_FILE):
    try:
        with open(CFG_FILE, "r", encoding="utf-8") as f:
            _saved_defaults = json.load(f)
    except Exception:
        pass

def _has_pil():
    try:
        import PIL; return True
    except ImportError:
        return False

THUMB_MAX = 100  # 缩略图批次上限
_undo_stack = []  # list of lists: [[(old,new),...], [(old,new),...]]
_window = None  # 存引用


def _err_human(e):
    """技术错误 → 人话（模块级，类方法内裸名可用）"""
    s = str(e).lower()
    if "timeout" in s or "timed out" in s: return "网络超时，请检查网络后重试"
    if "429" in s or "too many" in s: return "服务器繁忙，请稍后重试"
    if "所有" in str(e) and "不可达" in str(e): return "无法连接更新服务器，请检查网络"
    if "connection" in s or "refused" in s: return "网络不可达，请检查网络连接"
    return str(e)[:60]


class RenamerAPI:
    def pick_dest_folder(self):
        """打开文件夹选择框，返回路径"""
        result = _window.create_file_dialog(webview.FOLDER_DIALOG)
        if result and len(result) > 0:
            return {"path": result[0]}
        return {"path": ""}

    def generate_thumbnails(self, paths):
        """视频用 ffmpeg 抽帧，图片用 Pillow 缩略"""
        import subprocess, base64, tempfile, shutil, sys as _sys, json
        ffmpeg = shutil.which("ffmpeg")
        if not ffmpeg:
            for pfx in (_sys._MEIPASS if getattr(_sys,'_MEIPASS',False) else '', '/opt/homebrew/bin', '/usr/local/bin'):
                exe_name = 'ffmpeg.exe' if _sys.platform == 'win32' else 'ffmpeg'
                test = os.path.join(pfx, exe_name) if pfx else 'ffmpeg'
                if os.path.exists(test): ffmpeg = test; break
        # bundle 内 ffmpeg 优先
        meipass = getattr(_sys, '_MEIPASS', '')
        if meipass:
            bundled = os.path.join(meipass, 'ffmpeg')
            if os.path.exists(bundled): ffmpeg = bundled
        if not ffmpeg: ffmpeg = 'ffmpeg'
        IMG_EXT = {'.jpg','.jpeg','.png','.bmp','.tiff','.tif','.gif','.webp','.tga','.targa','.psd'}
        _log.info(f"generate_thumbnails: {len(paths)} files, ffmpeg={ffmpeg} PIL={'ok' if _has_pil() else 'MISSING'}")
        total = 0
        for p in paths[:THUMB_MAX]:
            try:
                ext = os.path.splitext(p)[1].lower()
                if ext in IMG_EXT:
                    if not _has_pil():
                        _log.warning(f"  thumb skip {os.path.basename(p)}: PIL not installed")
                        continue
                    from PIL import Image
                    img = Image.open(p)
                    # EXIF 自变换
                    try:
                        from PIL import ImageOps
                        img = ImageOps.exif_transpose(img)
                    except Exception:
                        pass
                    img.thumbnail((120, 120), Image.LANCZOS)
                    if img.mode in ('RGBA', 'P'):
                        img = img.convert('RGB')
                    from io import BytesIO
                    buf = BytesIO()
                    img.save(buf, format='JPEG', quality=80)
                    b64 = base64.b64encode(buf.getvalue()).decode()
                    thumb = f"data:image/jpeg;base64,{b64}"
                    if _window:
                        _window.evaluate_js(f"setThumb({json.dumps(p)},{json.dumps(thumb)})")
                    total += 1
                else:
                    # 视频：ffmpeg 抽帧
                    tmp = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                    tmp.close()
                    kw = dict(capture_output=True, timeout=8)
                    if _sys.platform == 'win32':
                        kw['creationflags'] = subprocess.CREATE_NO_WINDOW
                    for ss in ('00:00:01', '00:00:00.1'):
                        subprocess.run(
                            [ffmpeg, '-y', '-ss', ss, '-i', p, '-vframes', '1',
                             '-vf', 'scale=120:120:force_original_aspect_ratio=decrease',
                             '-q:v', '8', tmp.name],
                            **kw
                        )
                        if os.path.isfile(tmp.name) and os.path.getsize(tmp.name) > 100:
                            break
                    if os.path.isfile(tmp.name) and os.path.getsize(tmp.name) > 100:
                        with open(tmp.name, 'rb') as f:
                            b64 = base64.b64encode(f.read()).decode()
                        thumb = f"data:image/jpeg;base64,{b64}"
                        if _window:
                            _window.evaluate_js(f"setThumb({json.dumps(p)},{json.dumps(thumb)})")
                        total += 1
                    try: os.unlink(tmp.name)
                    except OSError: pass
            except Exception as e:
                _log.info(f"  thumb error: {e}")
                try: os.unlink(tmp.name)
                except OSError: pass
        _log.info(f"generate_thumbnails done: {total} thumbs")
        return {"thumbs": {}, "total": total}

    def get_config(self):
        fmt = []
        for fd in FIELD_CONFIG:
            nm = fd["name"]; k = fd["key"]
            if nm == "Ep": fmt.append({"pfx":"Ep","key":"ep"})
            elif nm == "Sc": fmt.append({"pfx":"Sc","key":"sc"})
            elif nm == "Gr": fmt.append({"pfx":"Gr","key":"gr"})
            elif nm == "Tk": fmt.append({"pfx":"Tk","key":"tk"})
            elif nm == "v": fmt.append({"pfx":"v","key":"ver"})
            elif k == "status": fmt.append({"pfx":"","key":"status"})
            else: fmt.append({"pfx":"","key":k})
        is_dev = not getattr(sys, '_MEIPASS', False)
        return {
            "fields": FIELD_CONFIG,  # 含 tk，P1a 动态 inspector 需要
            "dev": is_dev,
            "defaults": _saved_defaults,
            "method_desc_map": METHOD_DESC_MAP,
            "field_rules": FIELD_RULES,
            "name_format": fmt,
            "manual_url": _MANUAL_URL,
            "app_version": _APP_VERSION,
        }

    def open_manual(self):
        """使用手册：在线→浏览器打开，离线→返回 QR 码 base64"""
        import socket, base64
        from urllib.parse import urlparse
        try:
            host = urlparse(_MANUAL_URL).hostname
            s = socket.create_connection((host, 443), timeout=2)
            s.close()
            import webbrowser
            webbrowser.open(_MANUAL_URL)
            return {"ok": True, "method": "browser"}
        except Exception:
            pass
        try:
            from shared._qr import generate as _qr_generate
            matrix, size = _qr_generate(_MANUAL_URL.encode())
            scale = 4; pad = 12
            from io import BytesIO
            from PIL import Image
            img = Image.new("RGB", (size * scale + pad * 2, size * scale + pad * 2), "#ffffff")
            for r in range(size):
                for c in range(size):
                    if matrix[r][c]:
                        x = pad + c * scale
                        y = pad + r * scale
                        for dx in range(scale):
                            for dy in range(scale):
                                img.putpixel((x + dx, y + dy), (0, 0, 0))
            buf = BytesIO()
            img.save(buf, "PNG")
            b64 = base64.b64encode(buf.getvalue()).decode()
            return {"ok": True, "method": "qr", "qr": b64, "size": img.size[0], "url": _MANUAL_URL}
        except Exception as e:
            _log.warning(f"open_manual qr failed: {e}")
            return {"ok": False, "error": str(e)}

    def add_files_via_dialog(self):
        result = _window.create_file_dialog(
            webview.OPEN_DIALOG, allow_multiple=True,
            file_types=(_DIALOG_FILTER,),
        )
        if not result:
            return {"files": [], "total": 0}
        paths = result if isinstance(result, list) else [result]
        return self._process_paths(paths)

    def add_folder_via_dialog(self):
        result = _window.create_file_dialog(webview.FOLDER_DIALOG)
        if not result or len(result) == 0:
            return {"files": [], "total": 0}
        folder = result[0]
        paths = []
        try:
            for f in sorted(os.listdir(folder)):
                fp = os.path.join(folder, f)
                if os.path.isfile(fp):
                    paths.append(os.path.realpath(fp))
        except Exception:
            pass
        return self._process_paths(paths)

    def add_files_by_paths(self, paths):
        _log.info(f"add_files_by_paths paths={[str(p)[:80] for p in (paths or [])[:5]]}")
        return self._process_paths(paths)

    _dbg_buf = []  # 内存调试日志（最后 500 条）

    def debug_log(self, msg):
        if msg:
            self._dbg_buf.append(msg)
            if len(self._dbg_buf) > 500:
                self._dbg_buf = self._dbg_buf[-500:]
            # 立即写盘方便排错
            try:
                _real_stderr = getattr(sys, '__stderr__', sys.stderr)
                print(f"[JS] {msg}", file=_real_stderr)
            except Exception:
                pass
            _log.info(f"[JS] {msg}")
            return "ok"
        return {"log": list(self._dbg_buf)}

    def export_debug_package(self):
        """打包完整诊断信息 → 用户选择目录 → ZIP → Finder/Explorer 定位"""
        import zipfile, subprocess as _sp, socket, time, platform
        is_win = sys.platform == "win32"
        _CF = _sp.CREATE_NO_WINDOW if is_win else 0  # 隐藏控制台黑框

        # ── 选目录 ──
        dest = ""
        try:
            if is_win:
                ps_code = ('Add-Type -AssemblyName System.Windows.Forms; '
                           '$f = New-Object System.Windows.Forms.FolderBrowserDialog; '
                           '$f.Description = "选择导出位置"; $f.ShowDialog(); $f.SelectedPath')
                r = _sp.run(["powershell", "-NoProfile", "-Command", ps_code],
                            capture_output=True, text=True, timeout=120, creationflags=_CF)
                dest = r.stdout.strip() if r.returncode == 0 else ""
            else:
                r = _sp.run(
                    ["osascript", "-e",
                     'POSIX path of (choose folder with prompt "选择导出位置")'],
                    capture_output=True, text=True, encoding="utf-8", timeout=120)
                dest = r.stdout.strip()
        except Exception as e:
            _log.warning(f"export_debug: choose folder failed: {e}")
        if not dest or not os.path.isdir(dest):
            return {"ok": False, "error": "未选择目录"}

        # ── 日志目录 ──
        if is_win:
            _log_root = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")),
                                     "批量命名工具", "Logs")
        else:
            _log_root = os.path.join(os.path.expanduser("~"), "Library", "Logs", "批量命名工具")

        # ── ZIP 文件名 ──
        now = time.localtime()
        zip_name = f"批量命名工具-诊断日志-{now.tm_mon:02d}{now.tm_mday:02d}-{now.tm_hour:02d}{now.tm_min:02d}.zip"
        zip_path = os.path.join(dest, zip_name)

        def _add_str(zf, name, lines):
            zf.writestr(name, "\n".join(lines).encode("utf-8"))

        # ── info.txt ──
        info = []
        info.append(f"产品: 批量命名工具")
        info.append(f"版本: {_APP_VERSION}")
        info.append(f"产品ID: {_PRODUCT_ID}")
        info.append(f"系统: {platform.system()} {platform.release()}")
        info.append(f"Python: {sys.version}")
        info.append(f"主机名: {socket.gethostname()}")
        info.append(f"app路径: {getattr(sys, '_MEIPASS', 'N/A')}")
        info.append(f"时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")
        try:
            from shared.license import get_machine_fingerprint
            info.append(f"机器指纹: {get_machine_fingerprint()}")
        except Exception:
            info.append(f"机器指纹: N/A")
        info.append(f"设备模式: {'dev' if getattr(self, 'get_config', lambda: {})().get('dev') else 'prod'}")
        info.append(f"文件计数: {len(getattr(self, '_files', []))}")

        # ── state.txt ──
        state = []
        state.append(f"文件数: {len(getattr(self, '_files', []))}")
        state.append(f"Undo栈: {len(_undo_stack)}")
        state.append(f"更新状态: {json.dumps({k: str(v)[:100] for k,v in _UPDATE_STATE.items() if k != 'urls'}, ensure_ascii=False)}")
        # delta 覆盖目录状态
        delta_dir = os.path.expanduser('~/.config/renamer/delta')
        if os.path.isdir(delta_dir):
            state.append(f"delta目录: 存在")
            try:
                ver_file = os.path.join(delta_dir, 'version.txt')
                if os.path.isfile(ver_file):
                    with open(ver_file) as vf:
                        state.append(f"delta版本: {vf.read().strip()}")
            except Exception:
                pass
        else:
            state.append(f"delta目录: 不存在")

        # ── config.txt ──
        config_lines = []
        try:
            from shared import update_config as _uc
            config_lines.append(f"UPDATE_FILE: {_uc.UPDATE_FILE}")
            config_lines.append(f"TIMEOUT_VERSION_CHECK: {_uc.TIMEOUT_VERSION_CHECK}")
            for i, url in enumerate(_uc.DOWNLOAD_URLS):
                config_lines.append(f"DOWNLOAD_URL[{i}]: {url}")
        except Exception as e:
            config_lines.append(f"config读取失败: {e}")

        # ── network.txt ──
        net = []
        net.append(f"DNS: {socket.gethostbyname(socket.gethostname())}")
        try:
            r = _sp.run(["curl", "-sI", "--max-time", "8",
                "https://raw.githubusercontent.com/cgjpaladin/davinci-plugins/main/version.json"],
                capture_output=True, text=True, timeout=10)
            net.append(f"GitHub raw: HTTP {r.returncode} (stdout {len(r.stdout)}B)")
        except Exception as e:
            net.append(f"GitHub raw: 不可达 ({e})")
        try:
            r = _sp.run(["curl", "-sI", "--max-time", "8",
                "https://ghproxy.net/https://raw.githubusercontent.com/cgjpaladin/davinci-plugins/main/version.json"],
                capture_output=True, text=True, timeout=10)
            net.append(f"ghproxy: HTTP {r.returncode} (stdout {len(r.stdout)}B)")
        except Exception as e:
            net.append(f"ghproxy: 不可达 ({e})")

        # ── 写 ZIP ──
        try:
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                _add_str(zf, "info.txt", info)
                _add_str(zf, "state.txt", state)
                _add_str(zf, "config.txt", config_lines)
                _add_str(zf, "network.txt", net)
                # 内存调试日志
                if self._dbg_buf:
                    _add_str(zf, "debug_memory.log", self._dbg_buf)
                # Python 日志文件
                _log_file = os.path.join(_log_root, "renamer.log")
                if os.path.isfile(_log_file):
                    try:
                        zf.write(_log_file, "renamer.log")
                    except Exception:
                        pass
                # 当天/前一天日志
                if os.path.isdir(_log_root):
                    today = time.strftime("%Y-%m-%d")
                    yesterday = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
                    for f in sorted(os.listdir(_log_root)):
                        if (today in f or yesterday in f) and (f.endswith(".log") or f.endswith(".jsonl")):
                            try:
                                zf.write(os.path.join(_log_root, f), f"logs/{f}")
                            except Exception:
                                pass
                # 诊断日志文件
                _tmp = os.environ.get("TEMP") or os.environ.get("TMP") or "/tmp"
                for diag in ["apply_delta.log", "renamer_restart.log",
                             "bottle_route.log", "_renamer_restart.sh"]:
                    diag_path = os.path.join(_tmp, diag)
                    if os.path.isfile(diag_path):
                        try:
                            zf.write(diag_path, diag)
                        except Exception:
                            pass
            # 定位文件
            if is_win:
                _sp.run(["explorer", "/select,", zip_path], check=False)
            else:
                _sp.run(["open", "-R", zip_path], check=False)
            _log.info(f"export_debug: saved {zip_path}")
            return {"ok": True, "path": zip_path, "name": zip_name}
        except Exception as e:
            _log.warning(f"export_debug: zip failed: {e}")
            return {"ok": False, "error": _err_human(e)}

    def _process_paths(self, paths_):
        _log.info(f"_process_paths: {len(paths_)} paths → scanning")
        MAX_FILES = 100
        files = []; duplicates = 0; skipped = 0; subdirs = 0; truncated = False
        seen_fp = set()  # 本批指纹，不跨批
        _EMPTY_KEYS = {fd["key"] for fd in FIELD_CONFIG if fd.get("name")}  # 有前缀的字段不继承解析值
        parsed_count = 0; no_parse_count = 0
        for p_ in paths_[:MAX_FILES]:
            if len(files) >= MAX_FILES: truncated = True; break
            p = str(p_).strip()
            if p.startswith("file://"): p = unquote(p[7:])
            # smb:// → /Volumes/ 转换（Finder 拖入 SMB 文件时可能带 smb:// 前缀）
            p = re.sub(r'^smb://[\d.]+/', '/Volumes/', p)

            if os.path.isfile(p):
                ext = os.path.splitext(p)[1].lower()
                if ext not in SUPPORTED_EXT:
                    _log.debug(f"  skip non-video: {os.path.basename(p)}")
                    skipped += 1
                    continue
                if p in {f["path"] for f in files}: duplicates += 1; continue
                # 内容指纹：size + 前 64KB hash（同一文件走不同路径也能去重）
                try:
                    import hashlib
                    st = os.stat(p)
                    fp_key = f"{st.st_size}:{hashlib.md5(open(p,'rb').read(4096)).hexdigest()}"
                except OSError:
                    fp_key = p  # fallback to path
                if fp_key in seen_fp: duplicates += 1; continue
                seen_fp.add(fp_key)
                parsed = parse_filename(p)
                fields = {}
                for fd in FIELD_CONFIG:
                    k = fd["key"]
                    if parsed and k in parsed: fields[k] = parsed[k]
                    elif k in _EMPTY_KEYS: fields[k] = ""
                    else: fields[k] = fd["def"]
                if parsed: parsed_count += 1
                else: no_parse_count += 1
                _log.debug(f"  + {os.path.basename(p)} parsed={bool(parsed)} ep={fields.get('ep','')} sc={fields.get('sc','')} gr={fields.get('gr','')} desc={fields.get('desc','')}|method={fields.get('method','')}|author={fields.get('author','')}|v{fields.get('ver','')}|{fields.get('status','')}")
                files.append({"path":p,"basename":os.path.basename(p),"ext":os.path.splitext(os.path.basename(p))[1],"fields":fields,"fp":fp_key})
            elif os.path.isdir(p):
                try:
                    for f in sorted(os.listdir(p)):
                        fp = os.path.join(p, f)
                        if os.path.isfile(fp):
                            ext2 = os.path.splitext(fp)[1].lower()
                            if ext2 not in SUPPORTED_EXT: continue
                            if fp in {x["path"] for x in files}: duplicates += 1; continue
                            try:
                                st = os.stat(fp)
                                fp_key2 = f"{st.st_size}:{hashlib.md5(open(fp,'rb').read(4096)).hexdigest()}"
                            except OSError:
                                fp_key2 = fp
                            if fp_key2 in {x.get("fp","") for x in files}: duplicates += 1; continue
                            if len(files) >= MAX_FILES: truncated = True; break
                            parsed = parse_filename(fp)
                            fields = {}
                            for fd in FIELD_CONFIG:
                                k = fd["key"]
                                if parsed and k in parsed: fields[k] = parsed[k]
                                elif k in _EMPTY_KEYS: fields[k] = ""
                                else: fields[k] = fd["def"]
                            if parsed: parsed_count += 1
                            else: no_parse_count += 1
                            _log.debug(f"  + {os.path.basename(fp)} parsed={bool(parsed)} ep={fields.get('ep','')} sc={fields.get('sc','')}")
                            files.append({"path":fp,"basename":os.path.basename(fp),"ext":os.path.splitext(os.path.basename(fp))[1],"fields":fields,"fp":fp_key2})
                        elif os.path.isdir(fp): subdirs += 1
                except Exception: pass

        _log.info(f"_process_paths: {len(files)} files, {parsed_count} parsed, {no_parse_count} raw, {duplicates} dup, {skipped} skipped, {subdirs} subdirs, truncated={truncated}")
        files.sort(key=lambda f: f["basename"])
        # 自动检查
        anomalies = set()
        try:
            sp = [(fp, os.path.getsize(fp)) for f in files for fp in [f["path"]] if os.path.getsize(fp) > 0]
            vals = [s for _,s in sp]
            if len(vals) >= 3:
                mu = statistics.mean(vals); sd = statistics.stdev(vals)
                cv = sd/mu if mu > 0 else 0
                _log.info(f"auto-check: {len(vals)} files, mean={mu/1024/1024:.1f}MB sd={sd/1024/1024:.1f}MB cv={cv:.2f}")
                if cv > 1.0:
                    anomalies = {fp for fp,_ in sp if abs(os.path.getsize(fp)-mu) > 2*sd}
                    _log.info(f"  anomalies: {len(anomalies)}")
        except Exception as e:
            _log.info(f"  auto-check error: {e}")
        for f in files:
            tags = []
            if check_zero_byte(f["path"]): tags.append("zero")
            if check_double_ext(f["basename"]): tags.append("dbl_ext")
            if f["path"] in anomalies: tags.append("size")
            f["tags"] = tags
        return {"files":files,"total":len(files),"duplicates":duplicates,"skipped":skipped,"subdirs_skipped":subdirs,"truncated":truncated,"max":MAX_FILES}

    def do_rename(self, files):
        global _undo_stack
        ok = 0; fail = []; batch = []; renamed = []
        _log.info(f"do_rename: {len(files)} files")
        for f in files:
            p = f["path"]
            d = os.path.dirname(p)
            ext = os.path.splitext(os.path.basename(p))[1]
            nm = build_filename(f["fields"]) + ext
            np = os.path.join(d, nm)
            if os.path.exists(np) and os.path.normcase(np) != os.path.normcase(p):
                fail.append(os.path.basename(p) + " → 已存在")
                _log.warning(f"  rename collision: {os.path.basename(p)} → {os.path.basename(np)}")
                continue
            try:
                os.rename(p, np)
                batch.append((p, np))
                renamed.append({"old_path": p, "new_path": np})
                _log.debug(f"  ✓ {os.path.basename(p)} → {os.path.basename(np)}")
                ok += 1
            except Exception as e:
                fail.append(os.path.basename(p) + ": " + str(e))
        if files:
            sv = {k: v for k, v in files[0]["fields"].items() if k != "tk"}
            try:
                with open(CFG_FILE, "w", encoding="utf-8") as fp:
                    json.dump(sv, fp, ensure_ascii=False, indent=2)
                global _saved_defaults
                _saved_defaults = sv
            except Exception:
                pass
        if batch:
            _undo_stack.append({"type": "rename", "pairs": [(op, np) for op, np in batch]})
        _log.info(f"do_rename: {ok} ok, batch={len(batch)}, stack_depth={len(_undo_stack)}")
        return {"ok": ok, "fail": fail, "total": len(files), "renamed": renamed, "stack_depth": len(_undo_stack)}

    def do_undo(self):
        global _undo_stack
        if not _undo_stack:
            return {"ok": 0, "msg": "没有可撤销的操作"}
        entry = _undo_stack.pop()
        # 兼容旧格式: list of tuples
        if isinstance(entry, list):
            entry = {"type": "rename", "pairs": entry}
        typ = entry.get("type", "rename")
        pairs = entry.get("pairs", [])
        ud = 0; renamed = []
        for op, np in pairs:
            try:
                if typ == "archive":
                    os.remove(np)  # 删除已归档的文件
                    renamed.append({"old_path": np, "new_path": op})  # 通知 JS 还原文件引用
                else:
                    os.rename(np, op)
                    renamed.append({"old_path": np, "new_path": op})
                ud += 1
            except Exception:
                pass
        _log.info(f"do_undo: {ud}/{len(pairs)} {typ} reversed, remaining batches: {len(_undo_stack)}")
        return {"ok": ud, "msg": f"已撤销 {ud} 个", "remaining": len(_undo_stack), "renamed": renamed, "type": typ}

    def validate_dest(self, dest):
        v = str(dest).strip()
        if not v: return {"ok": False, "msg": ""}
        # smb:// → /Volumes/ 静默转换
        v = re.sub(r'^smb://[\d.]+/', '/Volumes/', v)
        # 检查路径有效性
        if os.path.isdir(v):
            return {"ok": True, "msg": "✓"}
        parent = os.path.dirname(v)
        if parent and os.path.isdir(parent):
            return {"ok": True, "msg": "✓ (将新建文件夹)"}
        return {"ok": False, "msg": "✗ 父目录不存在"}

    def do_archive(self, files, dest):
        import hashlib
        _log.info(f"do_archive: {len(files)} files, dest={dest}")
        ok = 0; fail = []; dup = 0; dest = os.path.realpath(re.sub(r'^smb://[\d.]+/', '/Volumes/', str(dest).strip()))
        def _hash_file(p):
            h = hashlib.sha256()
            with open(p, 'rb') as fh:
                while True:
                    chunk = fh.read(65536)
                    if not chunk: break
                    h.update(chunk)
            return h.digest()
        # 扫目标文件夹，预建 {hash: path}（同内容只保留一条，限 200 个文件，跳过 .tmp）
        seen = {}  # hash bytes → path
        if os.path.isdir(dest):
            for root, dirs, filenames in os.walk(dest):
                for fn in filenames:
                    if len(seen) >= 200: break
                    if fn.endswith('.tmp'): continue  # 跳过残片
                    fp = os.path.join(root, fn)
                    try:
                        sz = os.path.getsize(fp)
                        if sz > 0:
                            h = _hash_file(fp)
                            if h not in seen: seen[h] = fp
                    except OSError: pass
        # 逐文件处理
        archived = []
        for f in files:
            fd = f.get("fields", {})
            ext = f.get("ext", ".mp4")
            # 构建目标文件夹
            target = build_folder(dest, type('E',(),{'fields':fd,'ext':ext})())
            folder = os.path.dirname(target)
            os.makedirs(folder, exist_ok=True)
            # 扫描文件夹找最大 TK（同 Ep/Sc/Gr/desc/method/author/ver/status 归一组）
            try:
                max_tk = 0
                if os.path.isdir(folder):
                    # 用 tk='00' 构建模板，split 精准切分前缀
                    fd_copy = dict(fd); fd_copy['tk'] = '00'
                    sample = build_filename(fd_copy)
                    parts = sample.split('_Tk00_', 1)
                    if len(parts) == 2:
                        tk_prefix = parts[0] + '_Tk'
                        for fn in os.listdir(folder):
                            if fn.endswith('.tmp'): continue  # 跳过残片
                            if not fn.startswith(tk_prefix): continue
                            m = re.search(r'_Tk(0[1-9]|[1-9]\d)(?:_|\.mp4|\.mov|\.mxf|\.avi|\.mkv|$)', fn)
                            if m:
                                max_tk = max(max_tk, int(m.group(1)))
                nxt = max_tk + 1
                if nxt > 99: fail.append(f'{fd.get("ep","?")}_{fd.get("sc","?")}: TK 已满(99)'); continue
                fd['tk'] = str(nxt).zfill(2)
                target = os.path.join(folder, build_filename(fd) + ext)
            except Exception:
                fd['tk'] = '01'
                target = build_folder(dest, type('E',(),{'fields':fd,'ext':ext})())
            # 目标路径碰撞检测（同字段不同内容 → 会覆盖，不应发生）
            try:
                if os.path.exists(target):
                    # 尝试找下一个可用 TK
                    orig_tk = fd.get('tk', '01')
                    for n in range(int(orig_tk)+1, 100):
                        fd['tk'] = str(n).zfill(2)
                        alt = os.path.join(folder, build_filename(fd) + ext)
                        if not os.path.exists(alt):
                            target = alt; break
                    else:
                        fail.append(f'{os.path.basename(target)}: TK 已满(99)'); continue
                # 哈希去重
                src_hash = _hash_file(f["path"])
                if src_hash in seen: dup += 1; continue
                # 逐文件原子写入：copy2 → mark dedup → rename
                tmp_target = target + '.tmp'
                shutil.copy2(f["path"], tmp_target)
                seen[src_hash] = target  # 先标记去重——即使后续 move 失败，内容已在 .tmp
                shutil.move(tmp_target, target)
                archived.append((f["path"], target))
                ok += 1
                _log.debug(f"  ✓ {os.path.basename(f['path'])} → {os.path.basename(target)}")
            except Exception as e:
                fail.append(os.path.basename(f["path"]) + ": " + str(e))
                _log.warning(f"  ✗ {os.path.basename(f['path'])}: {e}")
        if archived:
            _undo_stack.append({"type": "archive", "pairs": archived})
        return {"ok": ok, "dup": dup, "fail": fail, "total": len(files), "archived": [{"old": src, "new": dst} for src, dst in archived], "stack_depth": len(_undo_stack)}

    def export_table(self, rows):
        """生成 xlsx 文件（openpyxl，含嵌入缩略图），返回 base64"""
        import base64 as _b64, io, re as _re
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from PIL import Image as PILImage

        # ── 列定义（单一事实来源，不再并行维护 HEADERS/KEYS/WIDTHS/NUM_KEYS）──
        COLUMNS = [
            {"header":"缩略图",  "key":"thumb",    "width":12, "zfill":False},
            {"header":"集数",    "key":"ep",       "width":5,  "zfill":True},
            {"header":"场次",    "key":"sc",       "width":5,  "zfill":True},
            {"header":"小场次",  "key":"gr",       "width":5,  "zfill":True},
            {"header":"次数",    "key":"tk",       "width":5,  "zfill":True},
            {"header":"描述",    "key":"desc",     "width":15, "zfill":False},
            {"header":"制作方式", "key":"method",   "width":10, "zfill":False},
            {"header":"制作者",   "key":"author",   "width":10, "zfill":False},
            {"header":"制作批次", "key":"ver",      "width":8,  "zfill":True},
            {"header":"通过情况", "key":"status",   "width":6,  "zfill":False},
            {"header":"文件名",   "key":"_newname", "width":45, "zfill":False},
        ]

        # ── 从 FIELD_CONFIG 推导（不手写第二份）──
        _PFX = {fd["key"]: fd["name"] for fd in FIELD_CONFIG if fd.get("name")}
        _FIELDS = [fd["key"] for fd in FIELD_CONFIG]

        # ── 布局常量 ──
        THUMB_W, THUMB_H_MAX = 72, 60
        DEFAULT_ROW_H = 15

        wb = Workbook()
        ws = wb.active
        ws.title = "文件列表"

        # 表头 + 列宽 + 冻结首行
        for ci, col in enumerate(COLUMNS):
            ws.cell(row=1, column=ci+1, value=col["header"])
            ws.column_dimensions[get_column_letter(ci+1)].width = col["width"]
        ws.freeze_panes = 'A2'

        # 数据行
        for rn, row in enumerate(rows):
            row_h = DEFAULT_ROW_H
            for ci, col in enumerate(COLUMNS):
                k = col["key"]
                if k == 'thumb':
                    thumb = row.get('thumb', '')
                    if thumb and thumb.startswith('data:image/'):
                        m = _re.match(r'data:image/(\w+);base64,(.+)', thumb)
                        if m:
                            data = _b64.b64decode(m.group(2))
                            try:
                                pil = PILImage.open(io.BytesIO(data))
                                pw, ph = pil.size
                                ratio = pw / ph if ph else 1
                                img = XLImage(io.BytesIO(data))
                                if ratio >= 1:
                                    img.width = THUMB_W
                                    img.height = max(1, int(THUMB_W / ratio))
                                else:
                                    img.height = THUMB_H_MAX
                                    img.width = max(1, int(THUMB_H_MAX * ratio))
                                import openpyxl.utils.units as oxu
                                row_pt = oxu.pixels_to_points(img.height) + 4
                                row_h = max(row_h, row_pt)
                            except Exception:
                                img = XLImage(io.BytesIO(data))
                                img.width = 60; img.height = 45
                                row_h = max(row_h, 50)
                            img.anchor = f'{get_column_letter(ci+1)}{rn+2}'
                            ws.add_image(img)
                elif k == '_newname':
                    parts = []
                    for fk in _FIELDS:
                        fv = str(row.get(fk, ''))
                        if fv:
                            pfx = _PFX.get(fk, '')
                            parts.append(f'{pfx}{fv}' if pfx else fv)
                    name = '_'.join(parts)
                    ws.cell(row=rn+2, column=ci+1, value=name + str(row.get('ext', '')))
                else:
                    val = str(row.get(k, ''))
                    if col.get("zfill") and val:
                        ws.cell(row=rn+2, column=ci+1, value=val.zfill(2))
                    else:
                        ws.cell(row=rn+2, column=ci+1, value=val)
            ws.row_dimensions[rn+2].height = row_h

        buf = io.BytesIO()
        wb.save(buf)
        return {'data': _b64.b64encode(buf.getvalue()).decode('ascii'), 'type': 'xlsx'}

    def save_file(self, data, default_name):
        """打开原生保存对话框，写入文件。返回 {ok: true/false, path: '...'}"""
        import base64 as _b64
        result = _window.create_file_dialog(
            webview.SAVE_DIALOG, save_filename=default_name,
            file_types=("Excel 文件 (*.xlsx)",),
        )
        if not result:
            return {"ok": False, "path": ""}
        try:
            with open(result, "wb") as f:
                f.write(_b64.b64decode(data))
            return {"ok": True, "path": result}
        except Exception as e:
            _log.warning(f"save_file failed: {e}")
            return {"ok": False, "path": result}

    # ═══ 自动更新（差分优先） ═══
    def check_update(self):
        global _UPDATE_STATE
        if "-dev" in _APP_VERSION:
            return {"update_available": False, "reason": "dev版不检查更新"}
        try:
            from shared import updater
            r = updater.check(_PRODUCT_ID, _APP_VERSION)
            if r.get("update_available"):
                _UPDATE_STATE["available"] = True
                _UPDATE_STATE["latest"] = r.get("latest", "")
                _UPDATE_STATE["notes"] = r.get("notes", "")
                _UPDATE_STATE["urls"] = r.get("urls", [])
            return r
        except Exception as e:
            _log.info(f"check_update error: {e}")
            return {"update_available": False, "reason": _err_human(e)}

    def trigger_bg_update(self):
        """后台查更新，不阻塞 JS 线程。完成后回调 JS onUpdateCheckDone()"""
        import threading, json as _json
        def _run():
            try:
                r = self.check_update()
                import webview
                windows = getattr(webview, 'windows', None)
                if windows and len(windows) > 0:
                    result_json = _json.dumps(r, ensure_ascii=False)
                    windows[0].evaluate_js(f'onUpdateCheckDone({result_json})')
            except Exception as e:
                _log.info(f"trigger_bg_update error: {e}")
                # 通知 JS 检测失败（json 序列化避免字符转义问题）
                try:
                    _reason = _err_human(e)
                    _reason_json = _json.dumps({"reason": _reason}, ensure_ascii=False)
                    import webview
                    windows = getattr(webview, 'windows', None)
                    if windows and len(windows) > 0:
                        windows[0].evaluate_js(f'onUpdateCheckDone({_reason_json})')
                except Exception:
                    pass
        threading.Thread(target=_run, daemon=True).start()

    def trigger_delta(self):
        """下载差分更新包（<200KB）"""
        global _UPDATE_STATE
        if not _UPDATE_STATE["available"]:
            return {"ok": False, "error": "没有可用更新"}
        if _UPDATE_STATE["downloading"]:
            return {"ok": False, "error": "已在下载中"}
        try:
            _UPDATE_STATE["downloading"] = True
            _UPDATE_STATE["downloaded"] = 0
            _UPDATE_STATE["total"] = 0

            import tempfile
            tmp_dir = tempfile.mkdtemp(prefix="renamer_delta_")
            zip_path = os.path.join(tmp_dir, "delta.zip")
            from shared import updater

            def _progress(dl, total):
                _UPDATE_STATE["downloaded"] = dl
                _UPDATE_STATE["total"] = total

            ok, err = updater.download_delta(zip_path, progress_callback=_progress)
            if not ok:
                _UPDATE_STATE["downloading"] = False
                return {"ok": False, "error": err}

            _UPDATE_STATE["zip_path"] = zip_path
            _UPDATE_STATE["ready"] = True
            _UPDATE_STATE["downloading"] = False
            _UPDATE_STATE["is_delta"] = True
            _log.info(f"delta update ready: {zip_path}")
            return {"ok": True}
        except Exception as e:
            _UPDATE_STATE["downloading"] = False
            _log.warning(f"trigger_delta error: {e}")
            return {"ok": False, "error": _err_human(e)}

    def apply_delta(self):
        """将差分文件写入当前 .app，重启应用。失败自动回滚。"""
        import tempfile
        global _UPDATE_STATE
        try: open(os.path.join(tempfile.gettempdir(), 'apply_delta.log'),'a').write(f"[{datetime.now():%H:%M:%S}] apply_delta called, ready={_UPDATE_STATE.get('ready')}, zip={'yes' if _UPDATE_STATE.get('zip_path') else 'no'}\n")
        except: pass
        if not _UPDATE_STATE.get("ready"):
            return {"ok": False, "error": "更新包未就绪"}
        try:
            meipass = getattr(sys, '_MEIPASS', '')
            if not meipass:
                return {"ok": False, "error": "无法定位 _MEIPASS"}
            # _MEIPASS = Contents/Frameworks（运行时不能写）→ 解压到 Contents/Resources
            fram_dir = meipass  # Contents/Frameworks/
            res_dir = os.path.join(os.path.dirname(fram_dir), 'Resources')  # Contents/Resources/
            if sys.platform == 'win32':
                app_path = os.path.dirname(sys.executable)
            else:
                app_path = os.path.dirname(os.path.dirname(meipass))
            if not app_path.endswith('.app'):
                p = meipass
                for _ in range(5):
                    if p.endswith('.app'): app_path = p; break
                    p = os.path.dirname(p)
            try: open(os.path.join(tempfile.gettempdir(), 'apply_delta.log'),'a').write(f"[{datetime.now():%H:%M:%S}] fram={fram_dir}, res={res_dir}, app={app_path}\n")
            except: pass

            import zipfile, tempfile, shutil as _sh, subprocess as _sp
            zip_path = _UPDATE_STATE["zip_path"]
            # macOS 禁止修改 .app bundle → 解压到 ~/.config/renamer/delta/
            delta_dir = os.path.expanduser('~/.config/renamer/delta')
            if os.path.exists(delta_dir):
                _sh.rmtree(delta_dir, ignore_errors=True)
            os.makedirs(delta_dir, exist_ok=True)
            try: open(os.path.join(tempfile.gettempdir(), 'apply_delta.log'),'a').write(f"[{datetime.now():%H:%M:%S}] extracting to {delta_dir}\n")
            except: pass
            with zipfile.ZipFile(zip_path, 'r') as zf:
                zf.extractall(delta_dir)
            try: open(os.path.join(tempfile.gettempdir(), 'apply_delta.log'),'a').write(f"[{datetime.now():%H:%M:%S}] extract done\n")
            except: pass

        except Exception as zip_err:
            try: open(os.path.join(tempfile.gettempdir(), 'apply_delta.log'),'a').write(f"[{datetime.now():%H:%M:%S}] ERROR: {zip_err}\n")
            except: pass
            return {"ok": False, "error": _err_human(zip_err)}

        # 解压后检查版本：delta 版本不能低于当前版本（防止旧代码覆盖新修复）
        try:
            _dv = os.path.join(delta_dir, 'version.txt')
            if os.path.isfile(_dv):
                with open(_dv, encoding='utf-8') as _f:
                    _dver = _f.read().strip()
                if _APP_VERSION:
                    try:
                        _dvt = tuple(int(x) for x in _dver.split('.'))
                        _avt = tuple(int(x) for x in _APP_VERSION.split('.'))
                        if _dvt < _avt:
                            import shutil
                            shutil.rmtree(delta_dir, ignore_errors=True)
                            return {"ok": False, "error": f"增量包版本过旧 ({_dver} < {_APP_VERSION})，已跳过"}
                    except ValueError:
                        pass
        except Exception:
            pass

        # 重启（launcher 自动从 ~/.config/renamer/delta/ 加载覆盖）
        _tmp = tempfile.gettempdir()
        is_win = sys.platform == "win32"
        if is_win:
            if os.path.isfile(os.path.join(app_path, "批量命名工具.exe")):
                binary = os.path.join(app_path, "批量命名工具.exe")
            elif os.path.isfile(os.path.join(app_path, "renamer_web.exe")):
                binary = os.path.join(app_path, "renamer_web.exe")
            else:
                binary = sys.executable  # --onefile fallback
                binary = os.path.join(app_path, "批量命名工具.exe")
        else:
            bundle_name = os.path.basename(app_path)
            if bundle_name.endswith('.app'):
                binary = os.path.join(app_path, 'Contents', 'MacOS', bundle_name[:-4])
            else:
                binary = os.path.join(app_path, 'Contents', 'MacOS', '批量命名工具')
            if not os.path.isfile(binary):
                binary = os.path.join(app_path, 'Contents', 'MacOS', '批量命名工具')
        if is_win:
            script_path = os.path.join(_tmp, '_renamer_restart.bat')
            restart_log = os.path.join(_tmp, 'renamer_restart.log')
            script = f'@echo off\nping -n 2 127.0.0.1 >nul\nstart "" "{binary}"\ndel "%~f0"\n'
        else:
            script_path = os.path.join(_tmp, '_renamer_restart.sh')
            restart_log = '/tmp/renamer_restart.log'
            script = (
                f'#!/bin/bash\nsleep 0.5\n'
                f'"{binary}" >> {restart_log} 2>&1\n'
                f'rm -f "$0"\n'
            )
        is_win = sys.platform == "win32"
        with open(script_path, 'w', encoding='gbk' if is_win else 'utf-8') as sf:
            sf.write(script)
        if not is_win:
            os.chmod(script_path, 0o755)
        try: open(os.path.join(tempfile.gettempdir(), 'apply_delta.log'),'a').write(f"[{datetime.now():%H:%M:%S}] launching restart via {script_path}\n")
        except: pass
        _sp.Popen(['/bin/bash', script_path] if sys.platform == 'darwin' else ['cmd', '/c', script_path],
            start_new_session=True, stdout=_sp.DEVNULL, stderr=_sp.DEVNULL,
            creationflags=_sp.CREATE_NO_WINDOW if is_win else 0)
        import time; time.sleep(0.2)
        os._exit(0)
        try:
            _UPDATE_STATE["downloading"] = True
            _UPDATE_STATE["downloaded"] = 0
            _UPDATE_STATE["total"] = 0

            # 创建临时下载目录
            import tempfile
            tmp_dir = tempfile.mkdtemp(prefix="renamer_update_")
            zip_path = os.path.join(tmp_dir, "update.zip")

            from shared import updater

            def _progress(dl, total):
                _UPDATE_STATE["downloaded"] = dl
                _UPDATE_STATE["total"] = total

            ok, err = updater.download_update(
                _PRODUCT_ID, zip_path, progress_callback=_progress)

            if not ok:
                _UPDATE_STATE["downloading"] = False
                return {"ok": False, "error": err}

            # 解压
            import zipfile
            extract_dir = os.path.join(tmp_dir, "extracted")
            os.makedirs(extract_dir, exist_ok=True)
            with zipfile.ZipFile(zip_path, 'r') as zf:
                zf.extractall(extract_dir)

            _UPDATE_STATE["zip_path"] = zip_path
            _UPDATE_STATE["extract_dir"] = extract_dir
            _UPDATE_STATE["ready"] = True
            _UPDATE_STATE["downloading"] = False
            _log.info(f"update ready: {extract_dir}")
            return {"ok": True}
        except Exception as e:
            _UPDATE_STATE["downloading"] = False
            _log.warning(f"trigger_update error: {e}")
            return {"ok": False, "error": str(e)[:100]}

    def apply_update(self):
        """写入更新脚本并 detach 执行，然后退出 app"""
        global _UPDATE_STATE
        if not _UPDATE_STATE["ready"]:
            return {"ok": False, "error": "更新包未就绪"}
        try:
            import tempfile, stat, subprocess

            # 找到当前 app 位置
            if sys.platform == 'darwin':
                meipass = getattr(sys, '_MEIPASS', '')
                _log.info(f"apply_update: _MEIPASS={meipass}")
                if not meipass:
                    return {"ok": False, "error": "无法定位 app 路径 (_MEIPASS 为空)"}
                app_path = os.path.dirname(os.path.dirname(meipass))  # .app bundle
            else:
                app_path = sys.executable  # Windows onefile: 直接取 exe 路径
            _log.info(f"apply_update: app_path={app_path}")

            # 新 app：从解压目录找第一个 .app 或 .exe
            ext_dir = _UPDATE_STATE["extract_dir"]
            _log.info(f"apply_update: ext_dir={ext_dir}, listing={os.listdir(ext_dir)[:5]}")
            new_app = None
            for item in os.listdir(ext_dir):
                full = os.path.join(ext_dir, item)
                if sys.platform == 'darwin' and item.endswith('.app'):
                    new_app = full; break
                elif sys.platform == 'win32' and item.endswith('.exe'):
                    new_app = full; break
            if not new_app:
                for root, dirs, files in os.walk(ext_dir):
                    for f in files + dirs:
                        if (sys.platform == 'darwin' and f.endswith('.app')) or \
                           (sys.platform == 'win32' and f.endswith('.exe')):
                            new_app = os.path.join(root, f); break
                    if new_app: break
            if not new_app:
                return {"ok": False, "error": "解压包中未找到 .app/.exe"}

            script_path = os.path.join(tempfile.gettempdir(),
                                       'renamer_update.command' if sys.platform == 'darwin' else 'renamer_update.bat')

            if sys.platform == 'darwin':
                need_admin = app_path.startswith('/Applications') or app_path.startswith('/Applications/')
                if need_admin:
                    script = f'''#!/bin/bash
sleep 2
xattr -d com.apple.quarantine "{new_app}" 2>/dev/null
osascript -e 'do shell script "rm -rf \\"{app_path}\\" && mv \\"{new_app}\\" \\"{app_path}\\" && xattr -d com.apple.quarantine \\"{app_path}\\" 2>/dev/null && open \\"{app_path}\\"" with administrator privileges'
rm -- "$0"
'''
                else:
                    script = f'''#!/bin/bash
sleep 2
rm -rf "{app_path}"
mv "{new_app}" "{app_path}"
xattr -d com.apple.quarantine "{app_path}" 2>/dev/null
open "{app_path}"
rm -- "$0"
'''
            else:  # Windows
                need_admin = 'Program Files' in app_path
                # 先 rename 旧 exe 防锁，再 move 新 exe，下次启动自动清 .old
                app_dir = os.path.dirname(app_path)
                app_name = os.path.basename(app_path)
                old_name = app_name + '.old'
                old_path = os.path.join(app_dir, old_name)
                if need_admin:
                    script = f'''@echo off
chcp 65001 >nul
timeout /t 2 /nobreak >nul
taskkill /f /im "{app_name}" 2>nul
ren "{app_path}" "{old_name}" 2>nul
powershell -Command "Start-Process cmd -ArgumentList '/c move /y \"\"\"{new_app}\"\"\" \"\"\"{app_path}\"\"\" && start \"\"\"\" \"\"\"{app_path}\"\"\" && del \"\"\"%~f0\"\"\"' -Verb RunAs"
'''
                else:
                    script = f'''@echo off
chcp 65001 >nul
timeout /t 2 /nobreak >nul
taskkill /f /im "{app_name}" 2>nul
ren "{app_path}" "{old_name}" 2>nul
move /y "{new_app}" "{app_path}" 2>nul
start "" "{app_path}"
del "{old_path}" 2>nul
del "%~f0"
'''

            with open(script_path, 'w', encoding='utf-8-sig' if sys.platform == 'win32' else 'utf-8') as f:
                f.write(script)
            os.chmod(script_path, 0o755)

            # detach 起脚本
            if sys.platform == 'darwin':
                _log.info(f"launching update script: {script_path}")
                if need_admin:
                    subprocess.Popen(
                        ['open', '-a', 'Terminal', script_path],
                        start_new_session=True,
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                else:
                    # 非管理员：nohup 后台运行，不弹终端窗
                    subprocess.Popen(
                        ['nohup', '/bin/bash', script_path],
                        start_new_session=True,
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:
                _log.info(f"launching update script: {script_path}")
                subprocess.Popen(
                    ['cmd', '/c', script_path],
                    creationflags=subprocess.CREATE_NO_WINDOW | subprocess.DETACHED_PROCESS,
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

            # 退出当前 app
            _window.destroy()
            os._exit(0)

        except Exception as e:
            _log.warning(f"apply_update error: {e}")
            return {"ok": False, "error": _err_human(e)}

    def get_update_progress(self):
        global _UPDATE_STATE
        return {
            "downloaded": _UPDATE_STATE["downloaded"],
            "total": _UPDATE_STATE["total"],
            "ready": _UPDATE_STATE["ready"],
            "available": _UPDATE_STATE["available"],
            "latest": _UPDATE_STATE["latest"],
        }

    def reveal_in_finder(self, path):
        """在 Finder 中显示文件"""
        import subprocess
        try:
            if sys.platform == 'darwin':
                subprocess.run(['open', '-R', path])
            elif sys.platform == 'win32':
                subprocess.run(['explorer', f'/select,{path}'])
        except Exception as e:
            _log.info(f"reveal_in_finder error: {e}")

    def get_media_info(self, path):
        """返回视频元数据（审查面板用），ffprobe 失败时降级为 ffmpeg"""
        import json, subprocess, shutil, sys as _sys
        try:
            ffprobe = shutil.which("ffprobe")
            if not ffprobe:
                for pfx in (_sys._MEIPASS if getattr(_sys,'_MEIPASS',False) else '', '/opt/homebrew/bin', '/usr/local/bin'):
                    exe = 'ffprobe.exe' if _sys.platform == 'win32' else 'ffprobe'
                    test = os.path.join(pfx, exe) if pfx else 'ffprobe'
                    if os.path.exists(test): ffprobe = test; break
            if not ffprobe: ffprobe = 'ffprobe'
            r = subprocess.run([ffprobe, '-v', 'quiet', '-print_format', 'json',
                '-show_format', '-show_streams', path],
                capture_output=True, text=True, timeout=5,
                creationflags=subprocess.CREATE_NO_WINDOW if _sys.platform=='win32' else 0)
            data = json.loads(r.stdout)
            streams = data.get('streams', [])
            fmt = data.get('format', {})
            vs = next((s for s in streams if s.get('codec_type') == 'video'), {})
            _as = next((s for s in streams if s.get('codec_type') == 'audio'), {})
            fps_str = vs.get('r_frame_rate', '0/1')
            try:
                n, d = fps_str.split('/')
                fps = float(n) / float(d) if float(d) != 0 else 0
            except (ValueError, ZeroDivisionError):
                fps = 0
            size_bytes = int(fmt.get('size', 0))
            return {
                'width': vs.get('width', 0), 'height': vs.get('height', 0),
                'duration': float(fmt.get('duration', 0)),
                'fps': round(fps, 2),
                'codec': vs.get('codec_name', ''),
                'size': size_bytes, 'size_mb': round(size_bytes / 1048576, 1),
                'bitrate_kbps': int(fmt.get('bit_rate', 0)) // 1000,
                'audio_codec': _as.get('codec_name', ''),
                'sample_rate': _as.get('sample_rate', ''),
            }
        except Exception:
            try:
                codec = self.get_codec(path)
                size_bytes = os.path.getsize(path) if os.path.exists(path) else 0
                return {
                    'width': 0, 'height': 0, 'duration': 0, 'fps': 0,
                    'codec': codec or '', 'size': size_bytes,
                    'size_mb': round(size_bytes / 1048576, 1),
                    'bitrate_kbps': 0, 'audio_codec': '', 'sample_rate': '',
                }
            except Exception:
                return None

    def get_codec(self, path):
        """用 ffmpeg 快速检测视频编码（跨平台备援）"""
        import subprocess, shutil, sys as _sys, re as _re
        try:
            ffmpeg = shutil.which("ffmpeg")
            if not ffmpeg:
                for pfx in (_sys._MEIPASS if getattr(_sys,'_MEIPASS',False) else '', '/opt/homebrew/bin', '/usr/local/bin'):
                    exe = 'ffmpeg.exe' if _sys.platform == 'win32' else 'ffmpeg'
                    test = os.path.join(pfx, exe) if pfx else 'ffmpeg'
                    if os.path.exists(test): ffmpeg = test; break
            meipass = getattr(_sys, '_MEIPASS', '')
            if meipass:
                bundled = os.path.join(meipass, 'ffmpeg')
                if os.path.exists(bundled): ffmpeg = bundled
            if not ffmpeg: ffmpeg = 'ffmpeg'
            r = subprocess.run([ffmpeg, '-i', path], capture_output=True, text=True, timeout=5,
                creationflags=subprocess.CREATE_NO_WINDOW if _sys.platform=='win32' else 0)
            m = _re.search(r'Video:\s*(\S+)', r.stderr)
            return m.group(1) if m else None
        except Exception:
            return None

    def get_media_data(self, path):
        """返回媒体文件 base64 + MIME（审查面板 Blob URL）"""
        import base64
        ext = os.path.splitext(path)[1].lower()
        mime_map = {'.mp4':'video/mp4','.mov':'video/quicktime','.avi':'video/x-msvideo','.mkv':'video/x-matroska','.webm':'video/webm','.mxf':'application/mxf','.m4v':'video/mp4','.flv':'video/x-flv','.png':'image/png','.jpg':'image/jpeg','.jpeg':'image/jpeg','.bmp':'image/bmp','.tiff':'image/tiff','.tif':'image/tiff','.gif':'image/gif','.webp':'image/webp','.tga':'image/x-targa','.targa':'image/x-targa','.psd':'image/vnd.adobe.photoshop'}
        try:
            if not os.path.isfile(path):
                _log.warning(f"get_media_data: file not found: {path}")
                return {'error': '文件不存在'}
            size = os.path.getsize(path)
            _log.info(f"get_media_data: {path} {size/1048576:.1f}MB")
            if size > 500 * 1024 * 1024:
                _log.warning(f"get_media_data: file too large ({size} bytes)")
                return {'error': f'文件过大 ({size/1048576:.0f}MB)，上限 500MB'}
            with open(path, 'rb') as f:
                data = f.read()
            return {'data': base64.b64encode(data).decode('ascii'), 'mime': mime_map.get(ext, 'application/octet-stream'), 'size': size}
        except Exception as e:
            _log.warning(f"get_media_data failed: {e}")
            return {'error': str(e)[:100]}

    def get_media_url(self, path):
        """返回可用于 <video src> 的 HTTP URL（绕过 WKWebView file:// 限制）"""
        import urllib.parse
        port = getattr(self, '_server_port', 0)
        if not port or not os.path.isfile(path):
            return ''
        return f'http://127.0.0.1:{port}/media?path={urllib.parse.quote(path, safe="")}'
_HTML_CANDIDATES = ["renamer_table.html", "renamer_web.html"]
HTML_FILE_NAME = "renamer_web.html"  # 默认
for _c in _HTML_CANDIDATES:
    if os.path.isfile(os.path.join(_BASE_DIR, _c)):
        HTML_FILE_NAME = _c; break
HTML_FILE = os.path.join(_BASE_DIR, HTML_FILE_NAME)

# ── 版本检测 ──
import re as _re_ver
_SCRIPT_JS = os.path.join(_BASE_DIR, 'app_table.js')
_HTML = os.path.join(_BASE_DIR, HTML_FILE_NAME) if os.path.isfile(os.path.join(_BASE_DIR, HTML_FILE_NAME)) else os.path.join(_BASE_DIR, 'renamer_table.html')
_APP_VERSION = '0.0.0'
_PRODUCT_ID = ('batch_renamer_mac' if sys.platform == 'darwin' else 'batch_renamer_win')
_MANUAL_URL = "https://jcnjno6i0upk.feishu.cn/docx/HEvydRFQZorKEnxBTYkconsinih?from=from_copylink"
try:
    # PyInstaller 打包后 app_table.js 不在，从拼接后的 HTML 中提取
    vfile = _HTML if os.path.isfile(_HTML) else _SCRIPT_JS
    if os.path.isfile(vfile):
        with open(vfile, 'r', encoding='utf-8') as _vf:
            content = _vf.read()
            _m = _re_ver.search(r"const APP_VERSION='([^']+)'", content)
            if _m: _APP_VERSION = _m.group(1)
    # 增量覆盖目录的版本文件（优先级最高，但不降级）
    _delta_ver = os.path.expanduser('~/.config/renamer/delta/version.txt')
    if os.path.isfile(_delta_ver):
        with open(_delta_ver, encoding='utf-8') as _dv:
            _dver = _dv.read().strip()
        # 如果 delta 版本低于内置版本，不使用 delta（清除覆盖目录）
        if _dver and _APP_VERSION:
            try:
                _dvt = tuple(int(x) for x in _dver.split('.'))
                _avt = tuple(int(x) for x in _APP_VERSION.split('.'))
                if _dvt < _avt:
                    import shutil
                    shutil.rmtree(os.path.dirname(_delta_ver), ignore_errors=True)
                    _dver = None
            except ValueError:
                pass
        if _dver:
            _APP_VERSION = _dver
except Exception:
    pass

# 更新状态（线程间共享）
_UPDATE_STATE = {
    "available": False, "latest": "", "notes": "", "urls": [],
    "downloading": False, "downloaded": 0, "total": 0, "ready": False,
    "zip_path": "", "extract_dir": "",
}

def _setup_native_menu_cocoa():
    """追加「检查更新」到 pywebview 现有 App 菜单。不替换，不重复。"""
    import subprocess, threading
    try:
        from Foundation import NSObject
        from AppKit import NSApplication, NSMenuItem
        from objc import selector
    except ImportError:
        return

    app = NSApplication.sharedApplication()
    main_menu = app.mainMenu()
    if not main_menu or main_menu.numberOfItems() == 0:
        return

    app_menu_item = main_menu.itemAtIndex_(0)
    app_menu = app_menu_item.submenu()
    if not app_menu:
        return

    # 持久化 receiver 列表（防止 GC 回收导致菜单灰掉）
    _receivers = getattr(_setup_native_menu_cocoa, '_receivers', None)
    if _receivers is None:
        _receivers = []
        _setup_native_menu_cocoa._receivers = _receivers

    # 复用 webview 的 update 模块：菜单回调触发 JS checkUpdate() → Python updater.check()
    class _CheckUpdateTarget(NSObject):
        @selector
        def checkForUpdates_(self, sender=None):
            import threading
            def _do():
                try:
                    import webview
                    windows = getattr(webview, 'windows', None)
                    if windows and len(windows) > 0:
                        windows[0].evaluate_js('checkUpdate()')
                except Exception:
                    pass
            threading.Thread(target=_do, daemon=True).start()

    target = _CheckUpdateTarget.alloc().init()
    target.retain()
    _receivers.append(target)

    item = NSMenuItem.alloc().initWithTitle_action_keyEquivalent_(
        '检查更新', 'checkForUpdates:', '')
    item.setTarget_(target)

    # 插入在分隔符之前（About 之后、Quit 之前）
    sep_idx = app_menu.indexOfItemWithTitle_('')  # 找第一个空标题项或遍历
    found = -1
    for i in range(app_menu.numberOfItems()):
        it = app_menu.itemAtIndex_(i)
        if it.isSeparatorItem():
            found = i
            break
    if found >= 0:
        app_menu.insertItem_atIndex_(item, found)
    else:
        app_menu.addItem_(item)


def main():
    """应用入口，由 launcher 调用"""
    import threading, socket
    from bottle import route, run, static_file

    # 用 bottle HTTP 服务绕过 WKWebView 沙箱限制
    _DELTA_HTML = os.path.expanduser('~/.config/renamer/delta')
    @route('/')
    def index():
        # 优先加载 delta 覆盖的 HTML
        if os.path.isfile(os.path.join(_DELTA_HTML, HTML_FILE_NAME)):
            try: open(os.path.join(tempfile.gettempdir(), 'bottle_route.log'),'w',encoding='utf-8').write(f"serving delta HTML, version={_APP_VERSION}\n")
            except: pass
            return static_file(HTML_FILE_NAME, root=_DELTA_HTML)
        try: open(os.path.join(tempfile.gettempdir(), 'bottle_route.log'),'w',encoding='utf-8').write(f"serving bundled HTML, version={_APP_VERSION}\n")
        except: pass
        return static_file(HTML_FILE_NAME, root=_BASE_DIR)

    @route('/media')
    def serve_media():
        """绕过 WKWebView 的 file:// 限制，通过 HTTP 提供视频/图片"""
        from bottle import request, Response, HTTPError
        path = request.query.path or request.query.get('path', '')
        if not path or not os.path.isfile(path):
            return HTTPError(404, "File not found")
        ext = os.path.splitext(path)[1].lower()
        mime_map = {'.mp4':'video/mp4','.mov':'video/quicktime','.mkv':'video/x-matroska','.webm':'video/webm',
                    '.avi':'video/x-msvideo','.png':'image/png','.jpg':'image/jpeg','.jpeg':'image/jpeg',
                    '.bmp':'image/bmp','.gif':'image/gif','.webp':'image/webp','.tiff':'image/tiff','.tif':'image/tiff',
                    '.tga':'image/x-targa','.targa':'image/x-targa','.psd':'image/vnd.adobe.photoshop'}
        return static_file(os.path.basename(path), root=os.path.dirname(path), mimetype=mime_map.get(ext, 'application/octet-stream'))

    # 找空闲端口
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.bind(('127.0.0.1', 0))
    port = sock.getsockname()[1]
    sock.close()

    # 后台启动 bottle
    t = threading.Thread(target=lambda: run(host='127.0.0.1', port=port, quiet=True), daemon=True)
    t.start()

    # 等 bottle 就绪
    import time, urllib.request
    for _ in range(20):
        try:
            urllib.request.urlopen(f"http://127.0.0.1:{port}", timeout=0.5)
            break
        except Exception:
            time.sleep(0.1)

    api = RenamerAPI()

    _window = webview.create_window(
        title="批量文件命名工具",
        url=f"http://127.0.0.1:{port}",
        js_api=api,
        width=880, height=620,
        min_size=(680, 400),
        resizable=True,
        background_color='#151515',
        text_select=True,
    )
    api._server_port = port  # 供 /media 路由使用

    # 拖放：用 loaded 事件在 DOM 就绪后绑定 Python 端 handler
    def _bind_drop():
        from webview.dom import DOMEventHandler
        def _on_drop(e):
            files = e['dataTransfer']['files']
            paths = []
            for f in files:
                fp = f.get('pywebviewFullPath', '')
                if not fp: continue
                if os.path.isfile(fp):
                    paths.append(os.path.realpath(fp))
                elif os.path.isdir(fp):
                    try:
                        for sf in sorted(os.listdir(fp)):
                            sfp = os.path.realpath(os.path.join(fp, sf))
                            if os.path.isfile(sfp):
                                paths.append(sfp)
                    except Exception: pass
            if not paths: return
            _log.info(f"DOM drop: {len(paths)} items [{', '.join(os.path.basename(p) for p in paths[:5])}{'...' if len(paths)>5 else ''}]")
            result = api._process_paths(paths)
            duplicates = result.get('duplicates', 0)
            if duplicates and not result.get('files'):
                _window.evaluate_js(f'toast("全部重复 · {duplicates} 个已跳过")')
            else:
                _window.evaluate_js(f"onDropResult({json.dumps(result)})")

        _window.dom.document.events.dragover += DOMEventHandler(lambda e: e, prevent_default=True)
        _window.dom.document.events.drop += DOMEventHandler(_on_drop, prevent_default=True, stop_propagation=True)
        _log.info("DOM drop handler bound")

    _window.events.loaded += _bind_drop

    # macOS 原生菜单——loaded 后在主线程设置
    if sys.platform == 'darwin':
        def _setup_menu_on_loaded():
            from Foundation import NSObject
            from objc import selector
            class _MenuSetupHelper(NSObject):
                @selector
                def performMenuSetup_(self, sender=None):
                    _setup_native_menu_cocoa()
            helper = _MenuSetupHelper.alloc().init()
            helper.performSelectorOnMainThread_withObject_waitUntilDone_(
                'performMenuSetup:', None, False)
        _window.events.loaded += _setup_menu_on_loaded

    webview.start(debug=False)
